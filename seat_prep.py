#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Oct  7 12:12:04 2020

@author: Piakman
"""
import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import pandas as pd

# Test Commands
def test_command():
    for i , row in blk_info_df.iterrows(): 
        print (i,row["Block"], row["Pivot"])
        xy = coordinate_from_string(row["Pivot"])
        cur_col = column_index_from_string(xy[0])
        cur_row = xy[1]
        print(xy, cur_col,cur_row)
        print('Value ',inside_sheet.cell(column=cur_col,row=cur_row).value)
    return

#blk_total = pd.DataFrame[['A',10,18],['B',10,18],['C',10,]]

def create_avail_seat():
    avail_seat = pd.DataFrame()
    i = 2
    block_seat_count = 0
    # Block Setup
    blck_name = blk_info_df.at[i,'Block']
    rownum = blk_info_df.at[i,'Row']
    seatnum = blk_info_df.at[i,'Seat']
    start_loc = blk_info_df.at[i,'Pivot']
    
    xy = coordinate_from_string(start_loc)
    cur_col = column_index_from_string(xy[0])
    cur_row = xy[1]
    

    if  blk_info_df.at[i,'Side']=='L':
        # Row Setup
        print('Left')
        row_step = -1
        end_row = cur_col - rownum-1
        row_count = 0
        #Seat Setup
        seat_step = -1
        end_seat = cur_row-seatnum-1
        
        for i in range(cur_col,end_row,row_step):
            row_count = row_count+1
            seat_count = 0
            for j in range(cur_row,end_seat,seat_step):
                seat_count=seat_count+1
                if inside_sheet.cell(column=i,row=j).value =='x':
                    block_seat_count = block_seat_count+1
                    print(block_seat_count,blck_name,row_count,seat_count,
                          inside_sheet.cell(column=i,row=j).value)
                    
        print(block_seat_count)
    elif blk_info_df.at[i,'Side']=='R':
        print('Right')
        # Row Setup
        row_step = 1
        end_row = cur_col + rownum+1
        row_count = 0
        #Seat Setup
        seat_step = -1
        end_seat = cur_row-seatnum-1

        for i in range(cur_col,end_row,row_step):
            row_count = row_count+1
            seat_count = 0
            for j in range(cur_row,end_seat,seat_step):
                seat_count=seat_count+1
                if inside_sheet.cell(column=i,row=j).value =='x':
                    block_seat_count = block_seat_count+1
                    print(block_seat_count,blck_name,row_count,seat_count,
                          inside_sheet.cell(column=i,row=j).value)
                    
        print(block_seat_count)
    elif blk_info_df.at[i,'Side']=='C':
        print('Center')
        # Row Setup
        row_step = 1
        end_row = cur_row + rownum
        row_count = 0
        #Seat Setup Since Center Exit is reverse then need to shift start cell
        cur_col = cur_col-seatnum+1
        seat_step = 1
        end_seat = cur_col+seatnum
        for i in range(cur_row,end_row,row_step):
            row_count = row_count+1
            seat_count = 0
            for j in range(cur_col,end_seat,seat_step):
                seat_count=seat_count+1
                print(i,get_column_letter(j))
                if inside_sheet.cell(column=j,row=i).value =='x':
                    block_seat_count = block_seat_count+1
                    print(block_seat_count,blck_name,row_count,seatnum-seat_count+1,
                          inside_sheet.cell(column=j,row=i).value)
                    
        print(block_seat_count)
    else :
        print('Special')
        row_step =-2
        
    return 

wb = openpyxl.load_workbook(filename = 'Config.xlsx')
blk_info_df =  pd.read_excel('Config.xlsx', sheet_name='Block Info')
sheet_list=wb.sheetnames
print(sheet_list)
inside_sheet=wb['Template Inside']

create_avail_seat()
 

    