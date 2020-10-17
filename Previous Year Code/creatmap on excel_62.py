# This code will setup seat in Commencement Ceremony
# Coded by: Yottana Khunatorn
# Jan 18, 2019

import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
from array import *

# Set cell border format
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
def getSeatmap(block):
    if block == "J":
        seatmap = [['n','n','n','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['c','s','s','s','s','s','s','s','s','s'],
                   ['c','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s']]

    elif block == "B":
        seatmap = [['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['c','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s']]

    elif block == "C":
        seatmap = [['n','n','n','n','n','n','n','n','s','s'],
                   ['n','n','n','n','n','n','s','s','s','s'],
                   ['n','n','n','n','s','s','s','s','s','s'],
                   ['n','n','s','s','s','s','s','s','s','s'],
                   ['n','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s']]        

    elif block == "D":
        seatmap = [['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','c','c','s','s','s','s']]
               
    elif block == "E":
        seatmap = [['s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s']]

    elif block == "F":
        seatmap = [['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','c','c','s','s','s','s','s','s','s','s','s','s','s','s']]

    elif block == "G":
        seatmap = [['s','s','n','n','n','n','n','n','n','n'],
                   ['s','s','s','s','n','n','n','n','n','n'],
                   ['s','s','s','s','s','s','n','n','n','n'],
                   ['s','s','s','s','s','s','s','s','n','n'],
                   ['s','s','s','s','s','s','s','s','s','n'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s']]        

    elif block == "H":
        seatmap = [['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','c'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s']]
        
    elif block == "I":
        seatmap = [['s','s','s','s','s','s','s','n','n','n'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','c'],
                   ['s','s','s','s','s','s','s','s','s','c'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s'],
                   ['s','s','s','s','s','s','s','s','s','s']]
        
    else:
        print("No block found")

    return(seatmap)        

def setULBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,Title):
    
    cur_cell = seatlayout.cell(row=PivotRow-1, column=PivotCol-1)
    for i in range(Totalcol,0,-1):
        cur_cell = seatlayout.cell(row=PivotRow-1, column=Totalcol+PivotCol-i)
        cur_cell.border = thin_border
        cur_cell.value = i
        cur_cell.font = Font(size = 8)
    for j in range(Totalrow,0,-1):
        cur_cell = seatlayout.cell(row=PivotRow+Totalrow-j, column=PivotCol-1)
        cur_cell.border = thin_border
        cur_cell.value = j
        cur_cell.font = Font(size = 8)

# Merge cell for Block Title
    seatlayout.merge_cells(start_row=PivotRow, start_column=PivotCol-2, end_row=PivotRow+Totalrow-1, end_column=PivotCol-2)
    cur_cell = seatlayout.cell(row=PivotRow, column=PivotCol-2)
    cur_cell.border = thin_border
    cur_cell.value = Title
    cur_cell.font = Font(size = 8,bold=True)
    cur_cell.alignment=Alignment(horizontal='general',
                     vertical='center',
                     text_rotation=90,
                     wrap_text=False)

    return

def setULBlockLayout(PivotRow,PivotCol,seatlayout,seatmap):

    currow = PivotRow
    for r in seatmap:
        curcol = PivotCol
        
        for c in r:
#            print currow,curcol
            cur_cell = seatlayout.cell(row=currow, column=curcol)
            if c =='n':
                print('No seat')
            elif c =='c':
                # Fill black color
                cur_cell.fill=PatternFill(bgColor='000000', fill_type = 'solid')
            elif c == 's':
                cur_cell.border = thin_border
            curcol = curcol+1
#        print ('Change row')
        currow = currow+1
    return

def setUCBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,Title):
    
    cur_cell = seatlayout.cell(row=PivotRow-1, column=PivotCol-1)
    for i in range(Totalcol,0,-1):
        cur_cell = seatlayout.cell(row=PivotRow-1, column=Totalcol+PivotCol-i)
        cur_cell.border = thin_border
        cur_cell.value = i
        cur_cell.font = Font(size = 8)
    for j in range(Totalrow,0,-1):
        cur_cell = seatlayout.cell(row=PivotRow+Totalrow-j, column=PivotCol-1)
        cur_cell.border = thin_border
        cur_cell.value = j
        cur_cell.font = Font(size = 8)

# Merge cell for Block Title
    seatlayout.merge_cells(start_row=PivotRow-2, start_column=PivotCol, end_row=PivotRow-2, end_column=PivotCol+Totalcol-1)
    cur_cell = seatlayout.cell(row=PivotRow-2, column=PivotCol)
    cur_cell.border = thin_border
    cur_cell.value = Title
    cur_cell.font = Font(size = 8,bold=True)
    cur_cell.alignment=Alignment(horizontal='center',
                     vertical='center',
                     text_rotation=0,
                     wrap_text=False)

    return

def setUCBlockLayout(PivotRow,PivotCol,seatlayout,seatmap):

    currow = PivotRow
    for r in seatmap:
        curcol = PivotCol
        
        for c in r:
#            print currow,curcol
            cur_cell = seatlayout.cell(row=currow, column=curcol)
            if c =='n':
                print('No seat')
            elif c =='c':
                # Fill black color
                cur_cell.fill=PatternFill(bgColor='000000', fill_type = 'solid')
            elif c == 's':
                cur_cell.border = thin_border
            curcol = curcol+1
#        print ('Change row')
        currow = currow+1
    return

def setURBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,Title):
    
    cur_cell = seatlayout.cell(row=PivotRow-1, column=PivotCol+1)
    for i in range(Totalcol,0,-1):
        cur_cell = seatlayout.cell(row=PivotRow-1, column= PivotCol-(Totalcol-i))
        cur_cell.border = thin_border
        cur_cell.value = i
        cur_cell.font = Font(size = 8)
    for j in range(Totalrow,0,-1):
        cur_cell = seatlayout.cell(row=PivotRow+Totalrow-j, column=PivotCol+1)
        cur_cell.border = thin_border
        cur_cell.value = j
        cur_cell.font = Font(size = 8)

# Merge cell for Block Title
    seatlayout.merge_cells(start_row=PivotRow, start_column=PivotCol+2, end_row=PivotRow+Totalrow-1, end_column=PivotCol+2)
    cur_cell = seatlayout.cell(row=PivotRow, column=PivotCol+2)
    cur_cell.border = thin_border
    cur_cell.value = Title
    cur_cell.font = Font(size = 8,bold=True)
    cur_cell.alignment=Alignment(horizontal='general',
                     vertical='center',
                     text_rotation=90,
                     wrap_text=False)

    return

def setURBlockLayout(PivotRow,PivotCol,Totalcol,seatlayout,seatmap):

    currow = PivotRow
    for r in seatmap:
        curcol = PivotCol-Totalcol+1
        
        for c in r:
#            print currow,curcol
            cur_cell = seatlayout.cell(row=currow, column=curcol)
            if c =='n':
                print('No seat')
            elif c =='c':
                # Fill black color
                cur_cell.fill=PatternFill(bgColor='000000', fill_type = 'solid')
            elif c == 's':
                cur_cell.border = thin_border
            curcol = curcol+1
#        print ('Change row')
        currow = currow+1
    return

def create_seatlayoutU():
# Create Sheet if it doesn't exist

    if ("U Seat Layout")in wb.sheetnames:
        print ('U Seat Layout Exist')
        seatlayout = wb.get_sheet_by_name("U Seat Layout")
        wb.remove_sheet(seatlayout)
        print ('Remove existing one .....')
        seatlayout = wb.create_sheet("U Seat Layout")
        print ('Create a new U Seat Layout .....')
    else:
        print ('Create a new U Seat Layout .....')
        seatlayout = wb.create_sheet("U Seat Layout")

#   Set Cell Size
    for col in range(1,151):
        i = get_column_letter(col)
        seatlayout.column_dimensions[i].width = float(4.6/2.12)    
        seatlayout.column_dimensions[i].font = Font( size = 7)

    for i in range(1,101):
       seatlayout.row_dimensions[i].height = float(3.7/0.35)



# create Block J seat map
    PivotRow = 2
    PivotCol = 3
    Totalrow = 19
    Totalcol = 10

    setULBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"J")
        
    seatmap = getSeatmap("J")

    setULBlockLayout(PivotRow,PivotCol,seatlayout,seatmap)            

# create Block B seat map
    PivotRow = 23
    PivotCol = 3
    Totalrow = 19
    Totalcol = 10

    setULBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"B")

    seatmap = getSeatmap("B")
    
    setULBlockLayout(PivotRow,PivotCol,seatlayout,seatmap)

# create Block C seat map
    PivotRow = 44
    PivotCol = 3
    Totalrow = 17
    Totalcol = 10

    setULBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"C")

    seatmap = getSeatmap("C")
    
    setULBlockLayout(PivotRow,PivotCol,seatlayout,seatmap)

# create Block D seat map
    PivotRow = 51
    PivotCol = 20
    Totalrow = 10
    Totalcol = 18

    setUCBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"D")

    seatmap = getSeatmap("D")
    
    setUCBlockLayout(PivotRow,PivotCol,seatlayout,seatmap)

# create Block E seat map
    PivotRow = 51
    PivotCol = 40
    Totalrow = 6
    Totalcol = 14
    
    setUCBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"E")

    seatmap = getSeatmap("E")
               
    setUCBlockLayout(PivotRow,PivotCol,seatlayout,seatmap)

# create Block F seat map
    PivotRow = 51
    PivotCol = 56
    Totalrow = 10
    Totalcol = 18

    setUCBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"F")

    seatmap = getSeatmap("F")
               
    setUCBlockLayout(PivotRow,PivotCol,seatlayout,seatmap)

# create Block G seat map
    PivotRow = 44
    PivotCol = 89
    Totalrow = 17
    Totalcol = 10

    setURBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"G")

    seatmap = getSeatmap("G")

    setURBlockLayout(PivotRow,PivotCol,Totalcol,seatlayout,seatmap)
    
# create Block H seat map
    PivotRow = 23
    PivotCol = 89
    Totalrow = 19
    Totalcol = 10

    setURBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"H")

    seatmap = getSeatmap("H")

    setURBlockLayout(PivotRow,PivotCol,Totalcol,seatlayout,seatmap)

# create Block I seat map
    PivotRow = 2
    PivotCol = 89
    Totalrow = 19
    Totalcol = 10

    setURBlockHeader(PivotRow,PivotCol,Totalrow,Totalcol,seatlayout,"I")

    seatmap = getSeatmap("I")

    setURBlockLayout(PivotRow,PivotCol,Totalcol,seatlayout,seatmap)
    
    return
    
def create_seatlayoutG():
    
# Create Sheet if it doesn't exist

    if ("G Seat Layout")in wb.sheetnames:
        print ('G Seat Layout Exist')
        seatlayout = wb.get_sheet_by_name("G Seat Layout")
        wb.remove_sheet(seatlayout)
        print ('Remove existing one .....')
        seatlayout = wb.create_sheet("G Seat Layout")
        print ('Create a new G Seat Layout.....')
    else:
        print ('Create a new G Seat Layout.....')
        seatlayout = wb.create_sheet("G Seat Layout")



# property cell.border should be used instead of cell.style.border
    cur_cell = seatlayout.cell(row=2, column=2)
    cur_cell.border = thin_border
    cur_cell.value = "Row"
    cur_cell.font = Font(size = 7)
    for i in range(3,43):
        cur_cell = seatlayout.cell(row=2, column=i)
        cur_cell.border = thin_border
        cur_cell.value = i-2
        cur_cell.font = Font(size = 7)
    for i in range(1,41):
        cur_cell = seatlayout.cell(row=2, column=43+i)
        cur_cell.border = thin_border
        cur_cell.value = i+40
        cur_cell.font = Font(size = 7)
    for i in range(1,34):
        cur_cell = seatlayout.cell(row=i+2, column=2)
        cur_cell.border = thin_border
        cur_cell.value = "A"+str(i)
        cur_cell.font = Font(size = 7)

# Create Seat Layout for Ground Floor
    RowNum = config_sheet.cell(row=2, column=2).value
    PivotRow = 2
    PivotCol = 2
    CurRow = 1
    print ('No. row of Block A: ',RowNum)

    cells = config_sheet['B6':'D38']
#    print cells

    for c1, c2,c3 in cells:
#        print("{0:8} {1:8} {1:8}".format(c1.value, c2.value,c3.value))
#        print (c1.value,c2.value,c3.value)
        SeatNum = c2.value
        Offset = c3.value
 
        for i in range(1,(SeatNum/2)+1):
            cur_cell = seatlayout.cell(row=PivotRow+CurRow, column=PivotCol+i+Offset)
            cur_cell.border = thin_border
            cur_cell.value = "x"
            cur_cell.font = Font(size = 7)
            cur_cell = seatlayout.cell(row=PivotRow+CurRow, column=PivotCol+(82-i-Offset))
            cur_cell.border = thin_border
            cur_cell.value = "x"
            cur_cell.font = Font(size = 7)
        CurRow=CurRow+1

# Create Seat Layout for Upper Floor without seat mark


# Set cell height and width and Fonts size

    seatlayout.column_dimensions['A'].width = float(3.88/2.12)
    seatlayout.column_dimensions['A'].font = Font( size = 7)
    seatlayout.column_dimensions['B'].width = float(5.64/2.12)
    seatlayout.column_dimensions['B'].font = Font( size = 7)
    for col in range(3,151):
        i = get_column_letter(col)
        seatlayout.column_dimensions[i].width = float(3.88/2.12)    
        seatlayout.column_dimensions[i].font = Font( size = 7)

    for i in range(1,101):
       seatlayout.row_dimensions[i].height = float(3.88/0.35)



    return





def mark_seat(sheet,c,r):
    thin_border = Border(
                        left=Side(border_style=BORDER_THIN, color='00000000'),
                        right=Side(border_style=BORDER_THIN, color='00000000'),
                        top=Side(border_style=BORDER_THIN, color='00000000'),
                        bottom=Side(border_style=BORDER_THIN, color='00000000')
                        )
    sheet.cell(row=11, column=10).border = thin_border
    print ('Mark Seat @',c,r)
    sheet.cell(row=c, column=r, value='x')
    return


wb = openpyxl.load_workbook(filename = 'Config.xlsx')
sheet_list=wb.sheetnames
print sheet_list
config_sheet= wb['Seat Config']
create_seatlayoutG()
create_seatlayoutU()
print(config_sheet['B6'].value)
wb.save('Config.xlsx')



