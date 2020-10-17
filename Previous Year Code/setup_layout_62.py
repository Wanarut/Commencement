#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Oct  4 11:47:53 2020

@author: Piakman
"""

import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.styles import *


def create_planlayout():
    # Create Sheet if it doesn't exist

    if ("G Seat Layout")in wb.sheetnames:
        print ('G Seat Layout Exist')
        seatlayout = wb.get_sheet_by_name("G Seat Layout")
        wb.remove_sheet(seatlayout)
        print ('Remove existing one .....')
        seatlayout = wb.create_sheet("G Seat Layout")
        print ('Create a new G Seat Layout.....')
    else:
        print ('Create a new G Seat Layout.....')
        seatlayout = wb.create_sheet("G Seat Layout")

    #   Set Cell Size
    for col in range(1,201):
        i = get_column_letter(col)
        seatlayout.column_dimensions[i].width = float(4.6/2.12)    
        seatlayout.column_dimensions[i].font = Font( size = 7)
        print(i)
    for i in range(1,151):
       seatlayout.row_dimensions[i].height = float(3.7/0.35)     
       print(i) 
    return


wb = openpyxl.load_workbook(filename = 'Config.xlsx')
sheet_list=wb.sheetnames
print (sheet_list)

create_planlayout()
