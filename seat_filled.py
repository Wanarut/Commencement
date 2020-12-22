#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 12 11:01:04 2020

@author: DAMASHII
"""
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill
import pandas as pd
import math
import numpy as np

first_reserved_size = 100
last_reserved_size = 332
catwalk_size = 4

print('Seat Step: ', end='')
s_seat_step = int(input())

if s_seat_step > catwalk_size:
    print('Must increase catwalk size')
    exit()

print('Morning or Afternoon [M/A]?', end='')
while(True):
    resp = input()
    if resp.upper() == 'M':
        out_template = 'Morning Order'
        break
    if resp.upper() == 'A':
        out_template = 'Afternoon Order'
        break
    print('Please response only M or A')

# out_template = 'Morning Order'
# out_template = 'Afternoon Order'

last_people_loc = []
color_i = 0
color_count = 0
reserveFill = PatternFill(fgColor='00FF00', fill_type='solid')
config_wb = openpyxl.load_workbook(filename='Config.xlsx')
list_wb = openpyxl.load_workbook(filename='List.xlsx')
row_no = 1
people_size = 1000


def main():
    block_info = pd.read_excel('Config.xlsx', sheet_name='Block Info')
    temp_inside = config_wb['Template Inside']
    print(block_info, '\n')

    temp_filled_morning = config_wb.copy_worksheet(temp_inside)
    temp_filled_morning.title = out_template

    list_wb.remove_sheet(list_wb[out_template+'_list1'])
    list_wb.create_sheet(out_template+'_list1')
    list_1 = list_wb[out_template+'_list1']
    list_1.cell(row=1, column=1).value = 'No.'
    list_1.cell(row=1, column=2).value = 'Block'
    list_1.cell(row=1, column=3).value = 'Line'
    list_1.cell(row=1, column=4).value = 'Seat'
    list_1.cell(row=1, column=5).value = 'Column'
    list_1.cell(row=1, column=6).value = 'Row'
    list_1.cell(row=1, column=7).value = 'Printable'

    list_wb.remove_sheet(list_wb[out_template+'_list2'])
    list_2 = list_wb.copy_worksheet(list_1)
    list_2.title = out_template+'_list2'

    blocks_seat_size = []
    for i in range(len(block_info)):
        blocks_seat_size.append(assign_available_block(block_info, i, temp_filled_morning, 'x'))
    
    print('Total available chairs are\t', sum(blocks_seat_size))
    global first_reserved_size
    first_reserved_size = sum(blocks_seat_size[:-1])
    print('Total available upper chairs are\t', first_reserved_size)
    print('Total available lowwer chairs are\t', blocks_seat_size[-1])
    fill_special_block(block_info, blocks_seat_size, temp_filled_morning)

    musical_chair(list_2, list_1)

    config_wb.save('Config.xlsx')
    list_wb.save('List.xlsx')


def assign_available_block(info, index, template, sign, people_size=0, p_info=None):
    block_seat_count = 0
    block = info.at[index, 'Block']
    line_size = info.at[index, 'Line']
    seat_size = info.at[index, 'Seat']
    MAX_seat = info.at[index, 'Max Seat']
    side = info.at[index, 'Side']
    lorder = info.at[index, 'L Order']
    sorder = info.at[index, 'S Order']
    pivot = info.at[index, 'Pivot']

    beg_loc = coordinate_from_string(pivot)
    beg_col = column_index_from_string(beg_loc[0])
    beg_row = beg_loc[1]

    interval = s_seat_step
    if side == 'L':
        line_step = -1
        seat_step = -s_seat_step
        if people_size == 0:
            print('Block', block, 'is Left Side', end='')
    elif side == 'R':
        line_step = 1
        seat_step = -s_seat_step
        if people_size == 0:
            print('Block', block, 'is Right Side', end='')
    elif side == 'C':
        line_step = 1
        seat_step = -s_seat_step
        if people_size == 0:
            print('Block', block, 'is Center Side', end='')
    elif side == 'S':
        line_step = 2
        seat_step = 1
        interval = 1
        seat_size = seat_size + catwalk_size
        if people_size == 0:
            print('Block', block, 'is Special Side', end='')
    else:
        if people_size == 0:
            print('Block', block, 'is N/A Side')
        return None

    if side == 'L' or side == 'R':
        line_size, seat_size = seat_size, line_size
        line_step, seat_step = seat_step, line_step
        lorder, sorder = sorder, lorder

    if lorder == 'r':
        if side == 'S':
            beg_row = beg_row + line_step*(line_size-1)
        else:
            beg_row = beg_row + np.sign(line_step)*(line_size-1)
        line_step = -line_step
        if people_size == 0:
            print('\tr line', end='')
    else:
        if people_size == 0:
            print('\tn line', end='')

    if sorder == 'r':
        beg_col = beg_col + np.sign(seat_step)*(seat_size-1)
        seat_step = -seat_step
        if people_size == 0:
            print('\tr seat')
    else:
        if people_size == 0:
            print('\tn seat')
    
    if side == 'S':
        end_row = beg_row + line_step*(line_size-1)
    else:
        end_row = beg_row + np.sign(line_step)*(line_size-1)
    end_col = beg_col + np.sign(seat_step)*(seat_size-1)
    
    if people_size == 0:
        print('Start\tat:', get_column_letter(beg_col), beg_row)
        print('End\tat:', get_column_letter(end_col), end_row)
    
    if side == 'L' or side == 'R':
        line_size, seat_size = seat_size, line_size
        # line_step, seat_step = seat_step, line_step
        lorder, sorder = sorder, lorder

    nextishead = False
    for i in range(line_size):
        seat_count = 0
        nextishead = True
        for j in range(int(seat_size/interval)+1):
            # rotation block
            if side == 'C' or side == 'S':
                cur_line, cur_seat = i, j
            else:
                cur_line, cur_seat = j, i

            if i%2:
                offset = 1
                if side == 'C':
                    col_offset = 1
                    row_offset = 0
                elif side != 'S':
                    col_offset = 0
                    row_offset = 1
            else:
                offset = 0
                col_offset = 0
                row_offset = 0

            cur_row = beg_row+(cur_line*line_step)+row_offset
            cur_col = beg_col+(cur_seat*seat_step)+col_offset
            
            # if i%2:
            #     if side == 'C':
            #         cur_col = beg_col+(cur_seat*seat_step)+1
            #     elif side != 'S':
            #         cur_row = beg_row+(cur_line*line_step)+1

            if side == 'S' and (cur_seat*seat_step) > (seat_size/2) - (catwalk_size/2):
                cur_col = cur_col + seat_step - 1

            list_1 = list_wb[out_template+'_list1']
            if template.cell(row=cur_row, column=cur_col).value == sign:
                seat_count = seat_count + 1
                block_seat_count = block_seat_count + 1
                if people_size > 0:
                    if sign == 'x':
                        template.cell(row=cur_row, column=cur_col).value = 'o'
                        template.cell(row=cur_row, column=cur_col).fill = reserveFill
                    else:
                        global color_i, color_count, row_no
                        template.cell(row=cur_row, column=cur_col).fill = PatternFill(fgColor=p_info.at[color_i, 'C_code'], fill_type='solid')
                        # template.cell(row=cur_row, column=cur_col).value = 'o'
                        template.cell(row=cur_row, column=cur_col).value = row_no
                        color_count = color_count + 1
                        row_no = row_no + 1
                        
                        list_1.cell(row=row_no, column=1).value = row_no - 1
                        list_1.cell(row=row_no, column=2).value = block
                        list_1.cell(row=row_no, column=5).value = get_column_letter(cur_col)
                        list_1.cell(row=row_no, column=6).value = cur_row
                        if j == int(seat_size/interval):
                            nextishead = True
                        if nextishead:
                            list_1.cell(row=row_no, column=7).value = 'Print'
                            nextishead = False

                        if side == 'S':
                            if lorder == 'n':
                                list_1.cell(row=row_no, column=3).value = i + 2
                            else:
                                list_1.cell(row=row_no, column=3).value = line_size - i + 1
                            if (cur_seat*seat_step) > (seat_size/2) - (catwalk_size/2):
                                if sorder == 'n':
                                    list_1.cell(row=row_no, column=4).value = ((j+1)*seat_step) - catwalk_size
                                else:
                                    list_1.cell(row=row_no, column=4).value = seat_size - (((j+1)*seat_step) - catwalk_size) - 1
                            else:
                                if sorder == 'n':
                                    list_1.cell(row=row_no, column=4).value = (j*seat_step) + 1
                                else:
                                    list_1.cell(row=row_no, column=4).value = seat_size - (j*seat_step)
                        
                        else:
                            if lorder == 'n':
                                list_1.cell(row=row_no, column=3).value = i + 1
                            else:
                                list_1.cell(row=row_no, column=3).value = line_size - i
                            if sorder == 'n':
                                list_1.cell(row=row_no, column=4).value = (j*s_seat_step) + 1 - offset
                            else:
                                list_1.cell(row=row_no, column=4).value = seat_size - (j*s_seat_step) - offset

                        if color_count == p_info.at[color_i, 'Size']:
                            color_i = color_i + 1
                            color_count = 0

                    if block_seat_count == people_size:
                        last_people_loc.append([cur_row, cur_col])
                        print('Block', block, 'End', get_column_letter(cur_col), cur_row, 'get people\t', people_size)
                        return
            else:
                if row_no > 1:
                    list_1.cell(row=row_no, column=7).value = 'Print'
                nextishead = True
                

        if people_size == 0:
            print('Line', block, i+1, '\thas', seat_count, '\tavailable chairs')
    
    if block_seat_count > MAX_seat:
        print('OVERFLOW Seat')

    if people_size == 0:
        print('available chairs are', block_seat_count, '\n')
        return block_seat_count


def fill_special_block(info, blocks_seat_size, template):
    s_loc = len(blocks_seat_size)-1

    # Import people
    p_info = import_people(blocks_seat_size)
    global people_size
    special_block_size = blocks_seat_size.pop(s_loc)
    remain_people_size = people_size - special_block_size

    info_copy = info
    if remain_people_size > 0:
        info = info.drop(s_loc)
        reserve_upper_block(info, blocks_seat_size, template, remain_people_size, p_info)
        assign_available_block(info_copy, s_loc, template, 'x', special_block_size, p_info)
        global config_wb
        config_wb.save('Config.xlsx')
        print('Please check seatable chair are \'o\' sign in Excel')
        print('continue [Y/n]?', end='')
        while(True):
            resp = input()
            if resp.upper() == 'Y':
                break
            if resp.upper() == 'N':
                exit()
            print('Please response only Y or N')
        config_wb = openpyxl.load_workbook(filename='Config.xlsx')
        template = config_wb[out_template]
        reorder_upper(info, blocks_seat_size, template, p_info)

        assign_available_block(info_copy, s_loc, template, 'o', special_block_size, p_info)
    else:
        assign_available_block(info_copy, s_loc, template, 'o', people_size, p_info,)


def reserve_upper_block(info, blocks_seat_size, template, people_size, p_info):
    print('Remaining to upper people are\t', people_size)

    mid_loc = int(len(blocks_seat_size)/2)
    mid_seat_size = blocks_seat_size[mid_loc]
    remain_people_size = people_size - mid_seat_size

    if remain_people_size < 0:
        assign_available_block(info, mid_loc, template, 'x', people_size, p_info)
        return

    assign_available_block(info, mid_loc, template, 'x', mid_seat_size, p_info)
    for block_loc in range(mid_loc):
        left_loc = mid_loc-block_loc-1
        right_loc = mid_loc+block_loc+1

        left_seat_size = blocks_seat_size[left_loc]
        right_seat_size = blocks_seat_size[right_loc]

        if remain_people_size < left_seat_size+right_seat_size:
            if remain_people_size % 2:
                # odd number of people
                assign_available_block(info, left_loc, template, 'x', int(remain_people_size/2)+1, p_info)
            else:
                # even number of people
                assign_available_block(info, left_loc, template, 'x', int(remain_people_size/2), p_info)

            assign_available_block(info, right_loc, template, 'x', int(remain_people_size/2), p_info)
            return

        remain_people_size = remain_people_size - left_seat_size - right_seat_size
        assign_available_block(info, left_loc, template, 'x', left_seat_size, p_info)
        assign_available_block(info, right_loc, template, 'x', right_seat_size, p_info)


def import_people(blocks_seat_size):
    p_info = pd.read_excel('Order.xlsx', sheet_name=out_template)
    global people_size
    people_size = sum(p_info['Size'])
    print('Total people are\t\t', people_size)
    if people_size > sum(blocks_seat_size):
        print('Number of total people are overflow, Must loop')
    return p_info


def reorder_upper(info, blocks_seat_size, template, p_info):
    global color_i, color_count, row_no
    color_i, color_count, row_no = 0, 0, 1

    for block_loc in range(len(info)-1, -1, -1):
        seat_size = blocks_seat_size[block_loc]
        assign_available_block(info, block_loc, template, 'o', seat_size, p_info)


def musical_chair(list_out, list_in):
    global people_size, row_no, config_wb, color_i, color_count
    color_i, color_count = 0, 0
    p_info = pd.read_excel('Order.xlsx', sheet_name=out_template)

    rotate_size = row_no - 1 - (first_reserved_size + last_reserved_size)
    remain_people = people_size-first_reserved_size-last_reserved_size
    loop_size = math.ceil(remain_people / rotate_size)
    print('Total Loop:', loop_size)

    template = []
    config_wb.remove_sheet(config_wb[out_template])
    for i in range(loop_size):
        temp_inside = config_wb['Template Inside']
        config_wb.remove_sheet(config_wb[out_template+'_'+str(i+1)])
        template.append(config_wb.copy_worksheet(temp_inside))
        template[i].title = out_template+'_'+str(i+1)

    for i in range(first_reserved_size):
        src_idx = i
        des_idx = i
        list_out.cell(row=des_idx+2, column=1).value = des_idx + 1
        for j in range(6):
            src_data = list_in.cell(row=src_idx+2, column=j+2).value
            if pd.notnull(src_data):
                list_out.cell(row=des_idx+2, column=j+2).value = src_data
        
        cur_row = list_out.cell(row=des_idx+2, column=6).value
        cur_col = column_index_from_string(list_out.cell(row=des_idx+2, column=5).value)
        for j in range(loop_size):
            template[j].cell(row=cur_row, column=cur_col).fill = PatternFill(fgColor=p_info.at[color_i, 'C_code'], fill_type='solid')
            template[j].cell(row=cur_row, column=cur_col).value = des_idx + 1
        # template[0].cell(row=cur_row, column=cur_col).fill = PatternFill(fgColor=p_info.at[color_i, 'C_code'], fill_type='solid')
        # template[0].cell(row=cur_row, column=cur_col).value = des_idx + 1
        color_count = color_count + 1
        if color_count == p_info.at[color_i, 'Size']:
            color_i = color_i + 1
            color_count = 0
            list_out.cell(row=des_idx+2, column=7).value = 'Print'
            list_out.cell(row=des_idx+3, column=7).value = 'Print'

    for i in range(remain_people):
        remainder = remain_people % rotate_size
        # src_idx = first_reserved_size + ((rotate_size-remainder+i) % rotate_size)
        # src_idx = first_reserved_size + (i % rotate_size)
        if int(i/rotate_size)+1 == loop_size:
            src_idx = first_reserved_size + ((rotate_size-remainder+i) % rotate_size)
        else:
            src_idx = first_reserved_size + (i % rotate_size)
        des_idx = first_reserved_size + i
        list_out.cell(row=des_idx+2, column=1).value = des_idx + 1
        for j in range(6):
            src_data = list_in.cell(row=src_idx+2, column=j+2).value
            if pd.notnull(src_data):
                list_out.cell(row=des_idx+2, column=j+2).value = src_data
            # list_out.cell(row=des_idx+2, column=j+2).value = list_in.cell(row=src_idx+2, column=j+2).value

        cur_row = list_out.cell(row=des_idx+2, column=6).value
        cur_col = column_index_from_string(list_out.cell(row=des_idx+2, column=5).value)
        j = int(i/rotate_size)
        template[j].cell(row=cur_row, column=cur_col).fill = PatternFill(fgColor=p_info.at[color_i, 'C_code'], fill_type='solid')
        template[j].cell(row=cur_row, column=cur_col).value = des_idx + 1
        color_count = color_count + 1
        if color_count == p_info.at[color_i, 'Size']:
            color_i = color_i + 1
            color_count = 0
            list_out.cell(row=des_idx+2, column=7).value = 'Print'
            list_out.cell(row=des_idx+3, column=7).value = 'Print'

    for i in range(last_reserved_size):
        src_idx = first_reserved_size + rotate_size + i
        des_idx = people_size - last_reserved_size + i
        list_out.cell(row=des_idx+2, column=1).value = des_idx + 1
        for j in range(6):
            src_data = list_in.cell(row=src_idx+2, column=j+2).value
            if pd.notnull(src_data):
                list_out.cell(row=des_idx+2, column=j+2).value = src_data
            # list_out.cell(row=des_idx+2, column=j+2).value = list_in.cell(row=src_idx+2, column=j+2).value
        
        cur_row = list_out.cell(row=des_idx+2, column=6).value
        cur_col = column_index_from_string(list_out.cell(row=des_idx+2, column=5).value)
        for j in range(loop_size):
            template[j].cell(row=cur_row, column=cur_col).fill = PatternFill(fgColor=p_info.at[color_i, 'C_code'], fill_type='solid')
            template[j].cell(row=cur_row, column=cur_col).value = des_idx + 1
        # template[0].cell(row=cur_row, column=cur_col).fill = PatternFill(fgColor=p_info.at[color_i, 'C_code'], fill_type='solid')
        # template[0].cell(row=cur_row, column=cur_col).value = des_idx + 1
        color_count = color_count + 1
        if color_count == p_info.at[color_i, 'Size']:
            color_i = color_i + 1
            color_count = 0
            list_out.cell(row=des_idx+2, column=7).value = 'Print'
            list_out.cell(row=des_idx+3, column=7).value = 'Print'


main()
